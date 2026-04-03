from __future__ import annotations

import argparse
import csv
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from email.utils import getaddresses
from html.parser import HTMLParser
from io import BytesIO, StringIO
import json
import os
from pathlib import Path
import shutil
import sys
import tempfile
from typing import Any, Literal
import zlib

from liteparse import CLINotFoundError, LiteParse, ParseError


SupportedFileType = Literal["pdf", "text", "email", "spreadsheet", "image", "unknown"]


PDF_MIME_TYPES = {"application/pdf"}
TEXT_MIME_TYPES = {
    "text/plain",
    "text/markdown",
}
EMAIL_MIME_TYPES = {"message/rfc822"}
SPREADSHEET_MIME_TYPES = {
    "text/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
IMAGE_MIME_TYPES = {
    "image/gif",
    "image/jpeg",
    "image/png",
    "image/webp",
}

TEXT_EXTENSIONS = {".txt", ".md", ".markdown"}
EMAIL_EXTENSIONS = {".eml", ".msg"}
SPREADSHEET_EXTENSIONS = {".csv", ".xls", ".xlsx"}
IMAGE_EXTENSIONS = {".gif", ".jpeg", ".jpg", ".png", ".webp"}


@dataclass(slots=True)
class DocumentSection:
    title: str
    text: str
    page: int | None = None
    index: int | None = None


@dataclass(slots=True)
class DocumentTable:
    name: str
    headers: list[str] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)


@dataclass(slots=True)
class DocumentAttachment:
    filename: str
    content_type: str | None = None
    size: int | None = None
    file_type: SupportedFileType = "unknown"
    disposition: str | None = None
    content_id: str | None = None
    is_inline: bool = False
    text: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)
    structured_data: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


@dataclass(slots=True)
class ParsedDocument:
    file_type: SupportedFileType
    text: str
    metadata: dict[str, Any] = field(default_factory=dict)
    sections: list[DocumentSection] = field(default_factory=list)
    tables: list[DocumentTable] = field(default_factory=list)
    attachments: list[DocumentAttachment] = field(default_factory=list)
    structured_data: dict[str, Any] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class ParserError(Exception):
    """Raised when the document cannot be parsed."""


def detect_file_type(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
) -> SupportedFileType:
    if isinstance(file_data, bytes):
        suffix = Path(filename or "").suffix.lower()
    else:
        suffix = Path(filename or file_data).suffix.lower()
    normalized_content_type = (content_type or "").split(";", 1)[0].strip().lower()

    if normalized_content_type in PDF_MIME_TYPES or suffix == ".pdf":
        return "pdf"

    if normalized_content_type in EMAIL_MIME_TYPES or suffix in EMAIL_EXTENSIONS:
        return "email"

    if normalized_content_type in SPREADSHEET_MIME_TYPES or suffix in SPREADSHEET_EXTENSIONS:
        return "spreadsheet"

    if normalized_content_type in IMAGE_MIME_TYPES or suffix in IMAGE_EXTENSIONS:
        return "image"

    if normalized_content_type in TEXT_MIME_TYPES or suffix in TEXT_EXTENSIONS:
        return "text"

    if isinstance(file_data, bytes):
        if file_data.startswith(b"%PDF-"):
            return "pdf"
        if _looks_like_png(file_data) or _looks_like_jpeg(file_data) or _looks_like_gif(file_data) or _looks_like_webp(file_data):
            return "image"
        if _looks_like_text(file_data):
            return "text"
        return "unknown"

    path = Path(file_data)
    if path.suffix:
        return "unknown"

    try:
        _decode_text(path.read_bytes())
    except (OSError, UnicodeDecodeError):
        return "unknown"
    return "text"


def parse_document(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
    timeout: float | None = 60.0,
    ocr_enabled: bool = True,
    ocr_language: str = "en",
    password: str | None = None,
) -> ParsedDocument:
    resolved_uploaded_at = _resolve_uploaded_at(uploaded_at)
    file_type = detect_file_type(file_data, filename=filename, content_type=content_type)

    if file_type == "pdf":
        return parse_pdf(
            file_data,
            filename=filename,
            content_type=content_type,
            uploaded_at=resolved_uploaded_at,
            timeout=timeout,
            ocr_enabled=ocr_enabled,
            ocr_language=ocr_language,
            password=password,
        )

    if file_type == "text":
        return parse_text(
            file_data,
            filename=filename,
            content_type=content_type,
            uploaded_at=resolved_uploaded_at,
        )

    if file_type == "email":
        return parse_email(
            file_data,
            filename=filename,
            content_type=content_type,
            uploaded_at=resolved_uploaded_at,
        )

    if file_type == "spreadsheet":
        return parse_spreadsheet(
            file_data,
            filename=filename,
            content_type=content_type,
            uploaded_at=resolved_uploaded_at,
        )

    if file_type == "image":
        return parse_image(
            file_data,
            filename=filename,
            content_type=content_type,
            uploaded_at=resolved_uploaded_at,
            timeout=timeout,
            ocr_enabled=ocr_enabled,
            ocr_language=ocr_language,
        )

    raise ParserError(
        "Unsupported file type. Supported types are PDF, plain text, email, spreadsheets, and common images."
    )


def parse_pdf(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
    timeout: float | None = 60.0,
    ocr_enabled: bool = True,
    ocr_language: str = "en",
    password: str | None = None,
    parser: LiteParse | None = None,
) -> ParsedDocument:
    resolved_uploaded_at = _resolve_uploaded_at(uploaded_at)
    liteparse_parser = parser or _build_liteparse_parser()
    warnings: list[str] = []

    try:
        result = liteparse_parser.parse(
            file_data,
            ocr_enabled=ocr_enabled,
            ocr_language=ocr_language,
            password=password,
            timeout=timeout,
        )
    except CLINotFoundError as exc:
        raise ParserError(
            "LiteParse CLI is not installed. Install Node.js >= 18 and run "
            "`npm install -g @llamaindex/liteparse`."
        ) from exc
    except ParseError as exc:
        raise ParserError(f"Failed to parse PDF: {exc}") from exc
    except FileNotFoundError as exc:
        raise ParserError(str(exc)) from exc

    if result.json is None:
        warnings.append("LiteParse did not return structured JSON for this PDF.")

    file_size = _get_file_size(file_data)
    sections = [
        DocumentSection(
            title=f"Page {page.pageNum}",
            text=page.text,
            page=page.pageNum,
            index=page.pageNum - 1,
        )
        for page in result.pages
    ]

    metadata: dict[str, Any] = {
        "filename": filename or _resolve_filename(file_data),
        "content_type": content_type or "application/pdf",
        "file_size": file_size,
        "uploaded_at": resolved_uploaded_at,
        "num_pages": result.num_pages,
        "ocr_enabled": ocr_enabled,
        "ocr_language": ocr_language,
    }

    if isinstance(file_data, (str, Path)):
        metadata["source_path"] = str(Path(file_data).resolve())

    return ParsedDocument(
        file_type="pdf",
        text=result.text,
        metadata=metadata,
        sections=sections,
        tables=[],
        attachments=[],
        structured_data=result.json or {},
        warnings=warnings,
    )


def parse_text(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
) -> ParsedDocument:
    resolved_uploaded_at = _resolve_uploaded_at(uploaded_at)
    raw_bytes = _read_bytes(file_data)
    text, encoding = _decode_text(raw_bytes)
    sections = _split_text_sections(text)

    metadata: dict[str, Any] = {
        "filename": filename or _resolve_filename(file_data),
        "content_type": content_type or "text/plain",
        "file_size": len(raw_bytes),
        "uploaded_at": resolved_uploaded_at,
        "encoding": encoding,
        "num_sections": len(sections),
    }

    if isinstance(file_data, (str, Path)):
        metadata["source_path"] = str(Path(file_data).resolve())

    return ParsedDocument(
        file_type="text",
        text=text,
        metadata=metadata,
        sections=sections,
        tables=[],
        attachments=[],
        structured_data={},
        warnings=[],
    )


def parse_email(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
) -> ParsedDocument:
    resolved_uploaded_at = _resolve_uploaded_at(uploaded_at)
    resolved_name = filename or _resolve_filename(file_data) or ""
    if Path(resolved_name).suffix.lower() == ".msg":
        raise NotImplementedError("Outlook .msg parsing is not implemented yet. Use .eml for now.")

    raw_bytes = _read_bytes(file_data)
    message = BytesParser(policy=policy.default).parsebytes(raw_bytes)

    plain_bodies: list[str] = []
    html_bodies: list[str] = []
    attachments: list[DocumentAttachment] = []
    warnings: list[str] = []

    for part in message.walk():
        if part.is_multipart():
            continue

        content_type_value = part.get_content_type()
        disposition = part.get_content_disposition()
        filename_value = part.get_filename()
        content_id = _normalize_content_id(part.get("Content-ID"))
        payload_bytes = _get_email_part_bytes(part)
        is_attachment = disposition == "attachment" or bool(filename_value)
        is_inline_attachment = disposition == "inline" and (
            bool(filename_value) or not content_type_value.startswith("text/")
        )

        if is_attachment or is_inline_attachment:
            attachments.append(
                _parse_email_attachment(
                    payload_bytes,
                    filename=filename_value or _default_attachment_name(content_type_value),
                    content_type=content_type_value,
                    disposition=disposition,
                    content_id=content_id,
                    uploaded_at=resolved_uploaded_at,
                )
            )
            continue

        if content_type_value == "text/plain":
            plain_text = _decode_email_text(payload_bytes, charset=part.get_content_charset())
            if plain_text.strip():
                plain_bodies.append(plain_text.strip())
            continue

        if content_type_value == "text/html":
            html_text = _decode_email_text(payload_bytes, charset=part.get_content_charset())
            rendered_html = _html_to_text(html_text)
            if rendered_html.strip():
                html_bodies.append(rendered_html.strip())

    body_text = "\n\n".join(plain_bodies) if plain_bodies else "\n\n".join(html_bodies)
    if not body_text.strip():
        warnings.append("The email did not contain a readable plain text or HTML body.")

    sections = _build_email_sections(
        subject=message.get("Subject", ""),
        plain_bodies=plain_bodies,
        html_bodies=html_bodies,
    )

    metadata: dict[str, Any] = {
        "filename": filename or _resolve_filename(file_data),
        "content_type": content_type or "message/rfc822",
        "file_size": len(raw_bytes),
        "uploaded_at": resolved_uploaded_at,
        "subject": message.get("Subject", ""),
        "from": _format_email_addresses(message.get_all("From", [])),
        "to": _format_email_addresses(message.get_all("To", [])),
        "cc": _format_email_addresses(message.get_all("Cc", [])),
        "bcc": _format_email_addresses(message.get_all("Bcc", [])),
        "reply_to": _format_email_addresses(message.get_all("Reply-To", [])),
        "date": message.get("Date", ""),
        "message_id": message.get("Message-ID", ""),
        "num_attachments": len(attachments),
    }

    if isinstance(file_data, (str, Path)):
        metadata["source_path"] = str(Path(file_data).resolve())

    warnings.extend(
        f"Attachment '{attachment.filename}': {warning}"
        for attachment in attachments
        for warning in attachment.warnings
    )

    return ParsedDocument(
        file_type="email",
        text=body_text,
        metadata=metadata,
        sections=sections,
        tables=[],
        attachments=attachments,
        structured_data={
            "headers": _extract_email_headers(message),
            "body": {
                "plain_text": plain_bodies,
                "html_text": html_bodies,
            },
        },
        warnings=warnings,
    )


def parse_image(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
    timeout: float | None = 60.0,
    ocr_enabled: bool = True,
    ocr_language: str = "en",
    parser: LiteParse | None = None,
) -> ParsedDocument:
    resolved_uploaded_at = _resolve_uploaded_at(uploaded_at)
    raw_bytes = _read_bytes(file_data)
    detected_content_type, image_format = _detect_image_kind(
        raw_bytes,
        filename=filename or _resolve_filename(file_data),
        content_type=content_type,
    )
    width, height = _extract_image_dimensions(raw_bytes, image_format)
    warnings: list[str] = []
    sections: list[DocumentSection] = []
    structured_data: dict[str, Any] = {}
    text = ""

    metadata: dict[str, Any] = {
        "filename": filename or _resolve_filename(file_data),
        "content_type": detected_content_type,
        "file_size": len(raw_bytes),
        "uploaded_at": resolved_uploaded_at,
        "image_format": image_format,
        "ocr_enabled": ocr_enabled,
        "ocr_language": ocr_language,
    }

    if width is not None and height is not None:
        metadata["width"] = width
        metadata["height"] = height

    if isinstance(file_data, (str, Path)):
        metadata["source_path"] = str(Path(file_data).resolve())

    description_parts = [f"Image attachment: {metadata['filename'] or 'unnamed image'}"]
    if image_format != "unknown":
        description_parts.append(f"Format: {image_format.upper()}")
    if width is not None and height is not None:
        description_parts.append(f"Dimensions: {width}x{height}")

    if ocr_enabled:
        liteparse_parser = parser or _build_liteparse_parser()
        try:
            ocr_source = _build_image_ocr_source(raw_bytes, image_format=image_format)
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as temp_pdf:
                temp_pdf.write(ocr_source)
                temp_pdf_path = Path(temp_pdf.name)
            try:
                result = liteparse_parser.parse(
                    temp_pdf_path,
                    ocr_enabled=True,
                    ocr_language=ocr_language,
                    timeout=timeout,
                )
            finally:
                temp_pdf_path.unlink(missing_ok=True)
        except CLINotFoundError:
            warnings.append(
                "LiteParse CLI is not installed. OCR could not run for this image."
            )
        except ParseError as exc:
            warnings.append(f"Image OCR failed: {exc}")
        except FileNotFoundError as exc:
            raise ParserError(str(exc)) from exc
        except ParserError as exc:
            warnings.append(str(exc))
        else:
            text = result.text.strip()
            sections = [
                DocumentSection(
                    title=f"Image Page {page.pageNum}",
                    text=page.text,
                    page=page.pageNum,
                    index=page.pageNum - 1,
                )
                for page in result.pages
            ]
            structured_data = result.json or {}
            metadata["num_pages"] = result.num_pages
            if result.json is None:
                warnings.append("LiteParse did not return structured JSON for this image.")

    if not text:
        text = "\n".join(description_parts)
    if not sections:
        sections = [DocumentSection(title="Image Metadata", text="\n".join(description_parts), index=0)]

    return ParsedDocument(
        file_type="image",
        text=text,
        metadata=metadata,
        sections=sections,
        tables=[],
        attachments=[],
        structured_data=structured_data,
        warnings=warnings,
    )


def parse_spreadsheet(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
) -> ParsedDocument:
    resolved_name = filename or _resolve_filename(file_data) or ""
    suffix = Path(resolved_name).suffix.lower()
    normalized_content_type = (content_type or "").split(";", 1)[0].strip().lower()

    if suffix == ".csv" or normalized_content_type == "text/csv":
        return _parse_csv_spreadsheet(
            file_data,
            filename=filename,
            content_type=content_type,
            uploaded_at=uploaded_at,
        )

    if suffix == ".xlsx" or normalized_content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _parse_xlsx_spreadsheet(
            file_data,
            filename=filename,
            content_type=content_type,
            uploaded_at=uploaded_at,
        )

    if suffix == ".xls" or normalized_content_type == "application/vnd.ms-excel":
        raise NotImplementedError(
            "Legacy .xls spreadsheet parsing is not implemented yet. Use .csv or .xlsx."
        )

    raise ParserError("Unsupported spreadsheet format. Use .csv or .xlsx.")


def parse_spreadsheets(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
) -> ParsedDocument:
    return parse_spreadsheet(
        file_data,
        filename=filename,
        content_type=content_type,
        uploaded_at=uploaded_at,
    )


def liteparse_runtime_available() -> bool:
    return _ensure_liteparse_environment() is not None


def save_parsed_document(document: ParsedDocument, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(document.to_dict(), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return output_path


def save_parsed_markdown(document: ParsedDocument, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    parts: list[str] = []
    title = document.metadata.get("filename") or "Parsed Document"
    parts.append(f"# {title}\n\n")

    if document.metadata:
        parts.append("## Metadata\n\n")
        for key, value in document.metadata.items():
            parts.append(f"- **{key}**: {value}\n")
        parts.append("\n")

    if document.sections:
        parts.append("## Sections\n\n")
        for section in document.sections:
            parts.append(f"### {section.title}\n\n")
            parts.append(f"{section.text}\n\n")
    else:
        parts.append("## Content\n\n")
        parts.append(f"{document.text}\n")

    if document.warnings:
        parts.append("\n## Warnings\n\n")
        for warning in document.warnings:
            parts.append(f"- {warning}\n")

    if document.attachments:
        parts.append("\n## Attachments\n\n")
        for attachment in document.attachments:
            parts.append(f"### {attachment.filename}\n\n")
            parts.append(f"- **content_type**: {attachment.content_type}\n")
            parts.append(f"- **file_type**: {attachment.file_type}\n")
            parts.append(f"- **size**: {attachment.size}\n")
            parts.append(f"- **disposition**: {attachment.disposition}\n")
            parts.append(f"- **is_inline**: {attachment.is_inline}\n")
            if attachment.content_id:
                parts.append(f"- **content_id**: {attachment.content_id}\n")
            for key, value in attachment.metadata.items():
                if key in {"filename", "content_type", "file_size"}:
                    continue
                parts.append(f"- **{key}**: {value}\n")
            parts.append("\n")

            if attachment.text:
                parts.append("#### Extracted Content\n\n")
                parts.append(f"{attachment.text}\n\n")

            if attachment.warnings:
                parts.append("#### Attachment Warnings\n\n")
                for warning in attachment.warnings:
                    parts.append(f"- {warning}\n")
                parts.append("\n")

    output_path.write_text("".join(parts), encoding="utf-8")
    return output_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Parse a document file and print the normalized JSON result."
    )
    parser.add_argument("file_path", help="Path to the file to parse.")
    parser.add_argument(
        "--content-type",
        help="Optional MIME type, such as application/pdf or text/plain.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=60.0,
        help="Timeout in seconds for PDF parsing. Default: 60.",
    )
    parser.add_argument(
        "--ocr-language",
        default="en",
        help="OCR language for PDF parsing. Default: en.",
    )
    parser.add_argument(
        "--no-ocr",
        action="store_true",
        help="Disable OCR when parsing PDFs.",
    )
    parser.add_argument(
        "--password",
        help="Password for protected PDFs.",
    )
    parser.add_argument(
        "--output",
        help="Optional output path. Use .json for structured output or .md/.txt for readable text output.",
    )

    args = parser.parse_args(argv)

    try:
        result = parse_document(
            args.file_path,
            content_type=args.content_type,
            timeout=args.timeout,
            ocr_enabled=not args.no_ocr,
            ocr_language=args.ocr_language,
            password=args.password,
        )
    except NotImplementedError as exc:
        print(f"Not implemented: {exc}", file=sys.stderr)
        return 1
    except ParserError as exc:
        print(f"Parser error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"Unexpected error: {exc}", file=sys.stderr)
        return 1

    output_json = json.dumps(result.to_dict(), indent=2, ensure_ascii=False)
    print(output_json)

    if args.output:
        output_path = Path(args.output)
        if output_path.suffix.lower() in {".md", ".markdown", ".txt"}:
            saved_path = save_parsed_markdown(result, output_path)
        else:
            saved_path = save_parsed_document(result, output_path)
        print(f"\nSaved parsed output to: {saved_path}", file=sys.stderr)

    return 0


def _build_liteparse_parser() -> LiteParse:
    cli_path = _ensure_liteparse_environment()
    return LiteParse(cli_path=cli_path, install_if_not_available=False)


def _parse_csv_spreadsheet(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
) -> ParsedDocument:
    resolved_uploaded_at = _resolve_uploaded_at(uploaded_at)
    raw_bytes = _read_bytes(file_data)
    text, encoding = _decode_text(raw_bytes)
    reader = csv.reader(StringIO(text))
    rows = [list(row) for row in reader]

    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    table = DocumentTable(name="Sheet1", headers=headers, rows=data_rows)
    sheet_text = _table_to_text(table)

    metadata: dict[str, Any] = {
        "filename": filename or _resolve_filename(file_data),
        "content_type": content_type or "text/csv",
        "file_size": len(raw_bytes),
        "uploaded_at": resolved_uploaded_at,
        "encoding": encoding,
        "num_sheets": 1,
        "sheet_names": ["Sheet1"],
        "num_rows": len(rows),
        "num_columns": len(headers),
    }

    if isinstance(file_data, (str, Path)):
        metadata["source_path"] = str(Path(file_data).resolve())

    return ParsedDocument(
        file_type="spreadsheet",
        text=sheet_text,
        metadata=metadata,
        sections=[DocumentSection(title="Sheet1", text=sheet_text, index=0)],
        tables=[table],
        attachments=[],
        structured_data={
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": headers,
                    "rows": data_rows,
                }
            ]
        },
        warnings=[],
    )


def _parse_xlsx_spreadsheet(
    file_data: str | Path | bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    uploaded_at: str | None = None,
) -> ParsedDocument:
    resolved_uploaded_at = _resolve_uploaded_at(uploaded_at)
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ParserError(
            "XLSX support requires openpyxl. Install project dependencies with uv sync."
        ) from exc

    workbook_source: str | BytesIO
    if isinstance(file_data, bytes):
        workbook_source = BytesIO(file_data)
        file_size = len(file_data)
    else:
        workbook_source = str(Path(file_data))
        file_size = Path(file_data).stat().st_size

    workbook = load_workbook(workbook_source, data_only=True, read_only=True)

    tables: list[DocumentTable] = []
    sections: list[DocumentSection] = []
    sheet_payload: list[dict[str, Any]] = []
    text_parts: list[str] = []

    for index, worksheet in enumerate(workbook.worksheets):
        rows = [
            [_normalize_spreadsheet_cell(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        rows = _trim_trailing_empty_rows(rows)
        if not rows:
            table = DocumentTable(name=worksheet.title, headers=[], rows=[])
            sheet_text = f"Sheet: {worksheet.title}\n\n(No data)"
        else:
            headers = [str(value) for value in rows[0]]
            data_rows = rows[1:] if len(rows) > 1 else []
            table = DocumentTable(name=worksheet.title, headers=headers, rows=data_rows)
            sheet_text = _table_to_text(table)

        tables.append(table)
        sections.append(DocumentSection(title=worksheet.title, text=sheet_text, index=index))
        text_parts.append(sheet_text)
        sheet_payload.append(
            {
                "name": worksheet.title,
                "headers": table.headers,
                "rows": table.rows,
            }
        )

    metadata: dict[str, Any] = {
        "filename": filename or _resolve_filename(file_data),
        "content_type": content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "file_size": file_size,
        "uploaded_at": resolved_uploaded_at,
        "num_sheets": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames,
    }

    if isinstance(file_data, (str, Path)):
        metadata["source_path"] = str(Path(file_data).resolve())

    workbook.close()

    return ParsedDocument(
        file_type="spreadsheet",
        text="\n\n".join(text_parts),
        metadata=metadata,
        sections=sections,
        tables=tables,
        attachments=[],
        structured_data={"sheets": sheet_payload},
        warnings=[],
    )


def _read_bytes(file_data: str | Path | bytes) -> bytes:
    if isinstance(file_data, bytes):
        return file_data
    return Path(file_data).read_bytes()


def _decode_text(raw_bytes: bytes) -> tuple[str, str]:
    encodings = ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "latin-1")
    last_error: UnicodeDecodeError | None = None

    for encoding in encodings:
        try:
            return raw_bytes.decode(encoding), encoding
        except UnicodeDecodeError as exc:
            last_error = exc

    if last_error is not None:
        raise last_error
    raise UnicodeDecodeError("utf-8", b"", 0, 1, "Unable to decode text file.")


def _looks_like_text(raw_bytes: bytes) -> bool:
    if not raw_bytes:
        return True

    sample = raw_bytes[:2048]
    if b"\x00" in sample:
        return False

    try:
        text = sample.decode("utf-8")
    except UnicodeDecodeError:
        return False

    printable_chars = sum(1 for char in text if char.isprintable() or char.isspace())
    return printable_chars / max(len(text), 1) > 0.9


def _looks_like_png(raw_bytes: bytes) -> bool:
    return raw_bytes.startswith(b"\x89PNG\r\n\x1a\n")


def _looks_like_jpeg(raw_bytes: bytes) -> bool:
    return raw_bytes.startswith(b"\xff\xd8\xff")


def _looks_like_gif(raw_bytes: bytes) -> bool:
    return raw_bytes.startswith((b"GIF87a", b"GIF89a"))


def _looks_like_webp(raw_bytes: bytes) -> bool:
    return len(raw_bytes) >= 12 and raw_bytes[:4] == b"RIFF" and raw_bytes[8:12] == b"WEBP"


def _split_text_sections(text: str) -> list[DocumentSection]:
    blocks = [block.strip() for block in text.split("\n\n") if block.strip()]
    if not blocks:
        return [DocumentSection(title="Body", text="", index=0)]

    return [
        DocumentSection(title=f"Section {index}", text=block, index=index - 1)
        for index, block in enumerate(blocks, start=1)
    ]


def _normalize_spreadsheet_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _trim_trailing_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    trimmed_rows = [list(row) for row in rows]
    while trimmed_rows and not any(cell for cell in trimmed_rows[-1]):
        trimmed_rows.pop()
    return trimmed_rows


def _table_to_text(table: DocumentTable) -> str:
    lines = [f"Sheet: {table.name}"]

    if table.headers:
        lines.append("")
        lines.append(" | ".join(table.headers))
        if table.rows:
            for row in table.rows:
                padded_row = row + [""] * max(0, len(table.headers) - len(row))
                lines.append(" | ".join(padded_row[: max(len(table.headers), len(padded_row))]))
    elif table.rows:
        lines.append("")
        for row in table.rows:
            lines.append(" | ".join(row))
    else:
        lines.append("")
        lines.append("(No data)")

    return "\n".join(lines)


def _get_file_size(file_data: str | Path | bytes) -> int:
    if isinstance(file_data, bytes):
        return len(file_data)
    return Path(file_data).stat().st_size


def _resolve_filename(file_data: str | Path | bytes) -> str | None:
    if isinstance(file_data, bytes):
        return None
    return Path(file_data).name


def _extract_email_headers(message: EmailMessage) -> dict[str, Any]:
    headers: dict[str, Any] = {}
    for key in message.keys():
        values = message.get_all(key, [])
        headers[key] = values if len(values) > 1 else (values[0] if values else "")
    return headers


def _format_email_addresses(values: list[str]) -> list[str]:
    return [
        f"{display_name} <{address}>" if display_name else address
        for display_name, address in getaddresses(values)
        if address
    ]


def _normalize_content_id(value: str | None) -> str | None:
    if value is None:
        return None
    return value.strip().strip("<>").strip() or None


def _get_email_part_bytes(part: EmailMessage) -> bytes:
    payload = part.get_payload(decode=True)
    if payload is not None:
        return payload

    text_payload = part.get_payload()
    if isinstance(text_payload, str):
        return text_payload.encode(part.get_content_charset() or "utf-8", errors="replace")
    return b""


def _decode_email_text(raw_bytes: bytes, *, charset: str | None = None) -> str:
    encodings = [charset] if charset else []
    encodings.extend(["utf-8", "utf-8-sig", "latin-1"])

    for encoding in encodings:
        if not encoding:
            continue
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue

    text, _ = _decode_text(raw_bytes)
    return text


class _HTMLToTextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in {"br", "div", "p", "li", "tr", "table", "section"}:
            self._parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"div", "p", "li", "tr", "table", "section"}:
            self._parts.append("\n")

    def handle_data(self, data: str) -> None:
        if data:
            self._parts.append(data)

    def get_text(self) -> str:
        text = "".join(self._parts)
        lines = [line.strip() for line in text.splitlines()]
        return "\n".join(line for line in lines if line)


def _html_to_text(html_text: str) -> str:
    parser = _HTMLToTextParser()
    parser.feed(html_text)
    parser.close()
    return parser.get_text()


def _build_email_sections(
    *,
    subject: str,
    plain_bodies: list[str],
    html_bodies: list[str],
) -> list[DocumentSection]:
    sections: list[DocumentSection] = []
    summary = f"Subject: {subject}" if subject else "Email Body"
    if plain_bodies:
        sections.append(
            DocumentSection(
                title="Plain Text Body",
                text="\n\n".join(plain_bodies),
                index=len(sections),
            )
        )
    if html_bodies:
        sections.append(
            DocumentSection(
                title="HTML Body",
                text="\n\n".join(html_bodies),
                index=len(sections),
            )
        )
    if not sections:
        sections.append(DocumentSection(title=summary, text="", index=0))
    return sections


def _default_attachment_name(content_type: str) -> str:
    subtype = content_type.split("/", 1)[-1] if "/" in content_type else "bin"
    return f"attachment.{subtype}"


def _parse_email_attachment(
    payload_bytes: bytes,
    *,
    filename: str,
    content_type: str,
    disposition: str | None,
    content_id: str | None,
    uploaded_at: str | None,
) -> DocumentAttachment:
    file_type = detect_file_type(payload_bytes, filename=filename, content_type=content_type)
    metadata: dict[str, Any] = {}
    structured_data: dict[str, Any] = {}
    extracted_text = ""
    warnings: list[str] = []

    try:
        parsed_attachment = parse_document(
            payload_bytes,
            filename=filename,
            content_type=content_type,
            uploaded_at=uploaded_at,
        )
    except NotImplementedError as exc:
        parsed_attachment = None
        warnings.append(str(exc))
    except ParserError as exc:
        parsed_attachment = None
        warnings.append(str(exc))
    except Exception as exc:
        parsed_attachment = None
        warnings.append(f"Unexpected attachment parsing error: {exc}")

    if parsed_attachment is not None:
        metadata = parsed_attachment.metadata
        structured_data = parsed_attachment.structured_data
        extracted_text = parsed_attachment.text
        warnings.extend(parsed_attachment.warnings)

    return DocumentAttachment(
        filename=filename,
        content_type=content_type,
        size=len(payload_bytes),
        file_type=file_type,
        disposition=disposition,
        content_id=content_id,
        is_inline=disposition == "inline",
        text=extracted_text,
        metadata=metadata,
        structured_data=structured_data,
        warnings=warnings,
    )


def _resolve_uploaded_at(uploaded_at: str | None) -> str:
    normalized = str(uploaded_at or "").strip()
    if normalized:
        return normalized
    return datetime.now(timezone.utc).isoformat()


def _detect_image_kind(
    raw_bytes: bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
) -> tuple[str, str]:
    normalized_content_type = (content_type or "").split(";", 1)[0].strip().lower()
    suffix = Path(filename or "").suffix.lower()

    if normalized_content_type in IMAGE_MIME_TYPES:
        return normalized_content_type, normalized_content_type.split("/", 1)[1]
    if suffix in IMAGE_EXTENSIONS:
        image_format = suffix.lstrip(".").replace("jpg", "jpeg")
        return f"image/{image_format}", image_format
    if _looks_like_png(raw_bytes):
        return "image/png", "png"
    if _looks_like_jpeg(raw_bytes):
        return "image/jpeg", "jpeg"
    if _looks_like_gif(raw_bytes):
        return "image/gif", "gif"
    if _looks_like_webp(raw_bytes):
        return "image/webp", "webp"
    return content_type or "application/octet-stream", "unknown"


def _extract_image_dimensions(raw_bytes: bytes, image_format: str) -> tuple[int | None, int | None]:
    if image_format == "png":
        if len(raw_bytes) >= 24:
            return int.from_bytes(raw_bytes[16:20], "big"), int.from_bytes(raw_bytes[20:24], "big")
        return None, None

    if image_format == "gif":
        if len(raw_bytes) >= 10:
            return int.from_bytes(raw_bytes[6:8], "little"), int.from_bytes(raw_bytes[8:10], "little")
        return None, None

    if image_format == "jpeg":
        return _extract_jpeg_dimensions(raw_bytes)

    return None, None


def _extract_jpeg_dimensions(raw_bytes: bytes) -> tuple[int | None, int | None]:
    width, height, _ = _extract_jpeg_info(raw_bytes)
    return width, height


def _extract_jpeg_info(raw_bytes: bytes) -> tuple[int | None, int | None, int | None]:
    if len(raw_bytes) < 4 or not _looks_like_jpeg(raw_bytes):
        return None, None, None

    offset = 2
    while offset + 9 < len(raw_bytes):
        if raw_bytes[offset] != 0xFF:
            offset += 1
            continue

        marker = raw_bytes[offset + 1]
        offset += 2

        if marker in {0xD8, 0xD9}:
            continue

        if offset + 2 > len(raw_bytes):
            return None, None, None

        segment_length = int.from_bytes(raw_bytes[offset : offset + 2], "big")
        if segment_length < 2 or offset + segment_length > len(raw_bytes):
            return None, None, None

        if marker in {
            0xC0, 0xC1, 0xC2, 0xC3,
            0xC5, 0xC6, 0xC7,
            0xC9, 0xCA, 0xCB,
            0xCD, 0xCE, 0xCF,
        }:
            if offset + 8 > len(raw_bytes):
                return None, None, None
            height = int.from_bytes(raw_bytes[offset + 3 : offset + 5], "big")
            width = int.from_bytes(raw_bytes[offset + 5 : offset + 7], "big")
            components = raw_bytes[offset + 7]
            return width, height, components

        offset += segment_length

    return None, None, None


def _ensure_liteparse_environment() -> str | None:
    cli_path = _find_liteparse_cli()
    if cli_path is None:
        return None

    bin_dir = str(Path(cli_path).parent)
    current_path = os.environ.get("PATH", "")
    path_entries = current_path.split(os.pathsep) if current_path else []
    if bin_dir not in path_entries:
        os.environ["PATH"] = os.pathsep.join([bin_dir, *path_entries]) if current_path else bin_dir

    return cli_path


def _find_liteparse_cli() -> str | None:
    explicit_path = os.environ.get("LITEPARSE_BIN")
    if explicit_path:
        resolved_explicit = _resolve_liteparse_candidate(explicit_path)
        if resolved_explicit is not None:
            return resolved_explicit

    direct_path = shutil.which("liteparse")
    if direct_path:
        return direct_path

    for candidate_dir in _candidate_liteparse_bin_dirs():
        resolved = _resolve_liteparse_candidate(Path(candidate_dir) / "liteparse")
        if resolved is not None:
            return resolved

    return None


def _candidate_liteparse_bin_dirs() -> list[Path]:
    candidates: list[Path] = []

    nvm_bin = os.environ.get("NVM_BIN")
    if nvm_bin:
        candidates.append(Path(nvm_bin).expanduser())

    nvm_dir = Path(os.environ.get("NVM_DIR", "~/.nvm")).expanduser()
    versions_dir = nvm_dir / "versions" / "node"
    if versions_dir.exists():
        version_dirs = sorted(
            (path for path in versions_dir.iterdir() if path.is_dir()),
            key=_node_version_sort_key,
            reverse=True,
        )
        candidates.extend(path / "bin" for path in version_dirs)

    return _dedupe_paths(candidates)


def _resolve_liteparse_candidate(candidate: str | Path) -> str | None:
    candidate_path = Path(candidate).expanduser()
    if not candidate_path.exists() or not candidate_path.is_file():
        return None
    if not os.access(candidate_path, os.X_OK):
        return None
    if candidate_path.is_absolute():
        return str(candidate_path)
    return str((Path.cwd() / candidate_path).absolute())


def _node_version_sort_key(path: Path) -> tuple[int, ...]:
    version_text = path.name.lstrip("v")
    parts: list[int] = []
    for part in version_text.split("."):
        try:
            parts.append(int(part))
        except ValueError:
            parts.append(-1)
    return tuple(parts)


def _dedupe_paths(paths: list[Path]) -> list[Path]:
    seen: set[str] = set()
    unique_paths: list[Path] = []
    for path in paths:
        resolved = str(path.expanduser())
        if resolved in seen:
            continue
        seen.add(resolved)
        unique_paths.append(path)
    return unique_paths


def _build_image_ocr_source(raw_bytes: bytes, *, image_format: str) -> bytes:
    if image_format == "png":
        width, height, rgb_bytes = _decode_png_to_rgb(raw_bytes)
        return _build_pdf_with_flate_image(width, height, rgb_bytes)

    if image_format == "jpeg":
        width, height, components = _extract_jpeg_info(raw_bytes)
        if width is None or height is None:
            raise ParserError("Could not read JPEG dimensions for OCR.")
        color_space = "DeviceGray" if components == 1 else "DeviceRGB"
        return _build_pdf_with_jpeg_image(width, height, raw_bytes, color_space=color_space)

    raise ParserError(
        f"OCR is only implemented for PNG and JPEG images in this version. Received: {image_format or 'unknown'}."
    )


def _decode_png_to_rgb(raw_bytes: bytes) -> tuple[int, int, bytes]:
    if not _looks_like_png(raw_bytes):
        raise ParserError("Unsupported PNG data for OCR.")

    offset = 8
    width: int | None = None
    height: int | None = None
    bit_depth: int | None = None
    color_type: int | None = None
    interlace_method: int | None = None
    idat_chunks: list[bytes] = []

    while offset + 8 <= len(raw_bytes):
        chunk_length = int.from_bytes(raw_bytes[offset : offset + 4], "big")
        chunk_type = raw_bytes[offset + 4 : offset + 8]
        chunk_data_start = offset + 8
        chunk_data_end = chunk_data_start + chunk_length
        if chunk_data_end + 4 > len(raw_bytes):
            raise ParserError("Corrupted PNG data.")

        chunk_data = raw_bytes[chunk_data_start:chunk_data_end]
        offset = chunk_data_end + 4

        if chunk_type == b"IHDR":
            width = int.from_bytes(chunk_data[0:4], "big")
            height = int.from_bytes(chunk_data[4:8], "big")
            bit_depth = chunk_data[8]
            color_type = chunk_data[9]
            interlace_method = chunk_data[12]
        elif chunk_type == b"IDAT":
            idat_chunks.append(chunk_data)
        elif chunk_type == b"IEND":
            break

    if width is None or height is None or bit_depth is None or color_type is None:
        raise ParserError("PNG header is incomplete.")
    if bit_depth != 8:
        raise ParserError("PNG OCR only supports 8-bit images in this version.")
    if interlace_method not in {0, None}:
        raise ParserError("Interlaced PNG OCR is not supported in this version.")
    if color_type not in {0, 2, 4, 6}:
        raise ParserError("PNG OCR only supports grayscale, RGB, grayscale+alpha, and RGBA images.")

    bytes_per_pixel = {
        0: 1,
        2: 3,
        4: 2,
        6: 4,
    }[color_type]

    decompressed = zlib.decompress(b"".join(idat_chunks))
    stride = width * bytes_per_pixel
    expected_length = (stride + 1) * height
    if len(decompressed) != expected_length:
        raise ParserError("Unexpected PNG pixel data length.")

    rgb_rows: list[bytes] = []
    previous_row = bytearray(stride)
    cursor = 0

    for _ in range(height):
        filter_type = decompressed[cursor]
        cursor += 1
        filtered = bytearray(decompressed[cursor : cursor + stride])
        cursor += stride
        row = _unfilter_png_scanline(filtered, previous_row, bytes_per_pixel, filter_type)
        rgb_rows.append(_png_scanline_to_rgb(bytes(row), color_type))
        previous_row = row

    return width, height, b"".join(rgb_rows)


def _unfilter_png_scanline(
    filtered: bytearray,
    previous_row: bytearray,
    bytes_per_pixel: int,
    filter_type: int,
) -> bytearray:
    row = bytearray(filtered)

    if filter_type == 0:
        return row

    if filter_type == 1:
        for index in range(len(row)):
            left = row[index - bytes_per_pixel] if index >= bytes_per_pixel else 0
            row[index] = (row[index] + left) & 0xFF
        return row

    if filter_type == 2:
        for index in range(len(row)):
            row[index] = (row[index] + previous_row[index]) & 0xFF
        return row

    if filter_type == 3:
        for index in range(len(row)):
            left = row[index - bytes_per_pixel] if index >= bytes_per_pixel else 0
            up = previous_row[index]
            row[index] = (row[index] + ((left + up) // 2)) & 0xFF
        return row

    if filter_type == 4:
        for index in range(len(row)):
            left = row[index - bytes_per_pixel] if index >= bytes_per_pixel else 0
            up = previous_row[index]
            up_left = previous_row[index - bytes_per_pixel] if index >= bytes_per_pixel else 0
            row[index] = (row[index] + _paeth_predictor(left, up, up_left)) & 0xFF
        return row

    raise ParserError(f"Unsupported PNG filter type: {filter_type}")


def _paeth_predictor(left: int, up: int, up_left: int) -> int:
    predictor = left + up - up_left
    left_distance = abs(predictor - left)
    up_distance = abs(predictor - up)
    up_left_distance = abs(predictor - up_left)

    if left_distance <= up_distance and left_distance <= up_left_distance:
        return left
    if up_distance <= up_left_distance:
        return up
    return up_left


def _png_scanline_to_rgb(scanline: bytes, color_type: int) -> bytes:
    if color_type == 2:
        return scanline

    if color_type == 6:
        return b"".join(scanline[index : index + 3] for index in range(0, len(scanline), 4))

    if color_type == 0:
        return b"".join(bytes((value, value, value)) for value in scanline)

    if color_type == 4:
        return b"".join(
            bytes((scanline[index], scanline[index], scanline[index]))
            for index in range(0, len(scanline), 2)
        )

    raise ParserError(f"Unsupported PNG color type: {color_type}")


def _build_pdf_with_flate_image(width: int, height: int, rgb_bytes: bytes) -> bytes:
    compressed = zlib.compress(rgb_bytes)
    image_dictionary = (
        f"<< /Type /XObject /Subtype /Image /Width {width} /Height {height} "
        f"/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /FlateDecode /Length {len(compressed)} >>"
    ).encode("ascii")
    return _build_pdf_document(
        width=width,
        height=height,
        image_stream=compressed,
        image_dictionary=image_dictionary,
    )


def _build_pdf_with_jpeg_image(width: int, height: int, jpeg_bytes: bytes, *, color_space: str) -> bytes:
    image_dictionary = (
        f"<< /Type /XObject /Subtype /Image /Width {width} /Height {height} "
        f"/ColorSpace /{color_space} /BitsPerComponent 8 /Filter /DCTDecode /Length {len(jpeg_bytes)} >>"
    ).encode("ascii")
    return _build_pdf_document(
        width=width,
        height=height,
        image_stream=jpeg_bytes,
        image_dictionary=image_dictionary,
    )


def _build_pdf_document(
    *,
    width: int,
    height: int,
    image_stream: bytes,
    image_dictionary: bytes,
) -> bytes:
    content_stream = f"q\n{width} 0 0 {height} 0 0 cm\n/Im0 Do\nQ\n".encode("ascii")
    content_dictionary = f"<< /Length {len(content_stream)} >>".encode("ascii")

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Count 1 /Kids [3 0 R] >>",
        (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {width} {height}] "
            f"/Resources << /XObject << /Im0 5 0 R >> >> /Contents 4 0 R >>"
        ).encode("ascii"),
        content_dictionary + b"\nstream\n" + content_stream + b"endstream",
        image_dictionary + b"\nstream\n" + image_stream + b"\nendstream",
    ]

    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{index} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")

    xref_offset = len(output)
    output.extend(f"xref\n0 {len(offsets)}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))

    trailer = f"trailer\n<< /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n"
    output.extend(trailer.encode("ascii"))
    return bytes(output)


if __name__ == "__main__":
    raise SystemExit(main())
